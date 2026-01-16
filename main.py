import pandas as pd
import os
from data_access.data_access_services import Data_access_services
from services import fichier_enquetes_to_agreg_sites_services as agreg_services
from services import prepa_data_services
from services import modeles_services_regression
from services import comparison_regression_models_services
from services import visu
from services.classification_models import classif_baselines
from services.classification_models import classif_models
from services.classification_models.adapters import RegressionToClassBaseline
from services import calcul_classes_services
from services import comparison_classification_models_services
from services import test_models
from services import cloud_visu

from matplotlib import pyplot as plt
import numpy as np
import math
import re
import sys


obj_data_access_services=Data_access_services()

pd.set_option('display.max_columns', None)
pd.set_option('display.max_rows', None)


nom_table_enquetes_agregees="enquetes_maisons"
schema= 'points_enquete'

LOAD_anf_AGGREG_DONNEES_ENQUETE_VF=False
LOAD_DONNEES_AGREGEES_VF=False
LOAD_DONNEES_BATIMENTS_VF=False
#Les données maisons->bâtiments->cosia sont calculées
LOAD_RAW_FEATURES_VF=True

TEST_TOUT_VF=False
    #_____________________
if TEST_TOUT_VF==False:
    TEST_BASES_LINES_VF=True
    #
    TEST_AGGREG_ABC_KNN_VF=True
    TEST_AGGREG_ABC_KNN_VISU_VF=False
    #
    TEST_VISU_COURBES_REGRESSION_VF=False
    #
    TEST_VISU_COURBES_AGREG_vf=False
    TEST_VISU_SCATTERS_par_GROUPE_v2_vf=False
    #
    TEST_APPROCHE_par_CLASSES_VF=False
    TEST_APPROCHE_GROUPEE_VF=False

else:
    TEST_BASES_LINES_VF=True
    TEST_AGGREG_ABC_KNN_UN_TYPE_VF=True
    #
    TEST_AGGREG_ABC_KNN_VF=True
    TEST_AGGREG_ABC_KNN_VISU_VF=True
    #
    TEST_VISU_COURBES_REGRESSION_VF=True 
    #
    TEST_VISU_SCATTERS_par_GROUPE_v2_vf=True 
    #
    TEST_APPROCHE_par_CLASSES_VF=True
    TEST_APPROCHE_GROUPEE_VF=True


    #CONTEXTE:
    if TEST_VISU_COURBES_REGRESSION_VF or TEST_AGGREG_ABC_KNN_VISU_VF:
            TEST_AGGREG_ABC_KNN_VF=True



if LOAD_anf_AGGREG_DONNEES_ENQUETE_VF:
    TEST_MULTI_VF=False
    AGREGATION_VF=True
    distance_max_pour_agreg_en_m=200.0
    EXPORT_TO_BD_VF=True
    agreg_services.enquetes_to_agreg_sites(TEST_MULTI_VF=TEST_MULTI_VF,AGREGATION_VF=AGREGATION_VF,distance_max_pour_agreg_en_m=distance_max_pour_agreg_en_m,EXPORT_TO_BD_VF=EXPORT_TO_BD_VF,nom_table_enquetes_agregees=nom_table_enquetes_agregees,schema=schema)

if LOAD_DONNEES_AGREGEES_VF:
    print(f'LOAD TABLE {nom_table_enquetes_agregees} from BD')
    gpd_maisons=obj_data_access_services.get_data_table_to_gdf(tablename=nom_table_enquetes_agregees)
    #
    print(f'\nTABLE {nom_table_enquetes_agregees} LOADED')
    print(f'->{len(gpd_maisons)} ligne(s) \n')
    print(gpd_maisons.head(1))
    #type_enquete  visite_VF
    print('\n')
    print(gpd_maisons['visite_VF'].value_counts())
    print('\n')
    print(gpd_maisons['type_enquete'].value_counts())
    #date_visite code_agent cluster_200m
    print('\n nb clusters')
    #print(gpd_maisons[['date_visite','code_agent','cluster_200m']].value_counts())
  

    freq_triplets = (
    gpd_maisons
    .value_counts(["date_visite", "code_agent", "cluster_200m"])
    .reset_index(name="count")
    )
    freq_distribution = (
        freq_triplets["count"]
        .value_counts()
        .sort_index()
        .reset_index(name="nb_triplets")
        .rename(columns={"index": "nb_lignes_par_triplet"})
    )
    #print(freq_distribution)

   
    maxi=15
    counts_left = freq_triplets.loc[
        freq_triplets["count"] <= maxi, "count"
    ]
    title_s=f"Fréquence du nombre d'habitat étudié par plainte (count ≤ {maxi})" 
    bins = np.arange(1, maxi+2)  # 1 à 21 → couvre 1..20

    plt.figure()
    plt.hist(counts_left, bins=bins, align="left")
    plt.xlabel("Nombre de lignes par triplet")
    plt.ylabel("Nombre de triplets")
    plt.title(title_s)
    plt.grid(axis="y")
    plt.show()

    #type_enquete
    # plainte           5187
    # porte_a_porte     2852
    # enquete_entomo     456

    freq_triplets = (
    gpd_maisons
    .value_counts(["date_visite", "code_agent", "cluster_200m","type_enquete"])
    .reset_index(name="count")
    )
    freq_distribution = (
        freq_triplets["count"]
        .value_counts()
        .sort_index()
        .reset_index(name="nb_triplets")
        .rename(columns={"index": "nb_lignes_par_triplet"})
    )

    df_left = freq_triplets[freq_triplets["count"] <= maxi]
    for type_enquete, sub in df_left.groupby("type_enquete"):
        plt.hist(
            sub["count"],
            bins=bins,
            align="left",
            alpha=0.5,
            label=type_enquete
        )

    plt.xlabel("Nombre de lignes par triplet")
    plt.ylabel("Nombre de triplets")
    plt.title(title_s)
    plt.legend()
    plt.grid(axis="y")
    plt.show()

    for type_enquete, sub in df_left.groupby("type_enquete"):
        plt.hist(
            sub["count"],
            bins=bins,
            align="left",
            alpha=0.5,
            density=True,
            label=type_enquete
        )

    plt.xlabel("Nombre de lignes par triplet")
    plt.ylabel("Fréquence relative")
    plt.title(title_s)
    plt.legend()
    plt.grid(axis="y")
    plt.show()


    df_left = freq_triplets[freq_triplets["count"] <= maxi]

    freq_by_type = (
        df_left
        .groupby(["count", "type_enquete"])
        .size()
        .unstack(fill_value=0)
        .sort_index()
    )
    x = freq_by_type.index.values
    types = freq_by_type.columns
    n_types = len(types)

    width = 0.25  # largeur des barres
    offsets = np.linspace(
        -width * (n_types - 1) / 2,
        width * (n_types - 1) / 2,
        n_types
    )

   
    freq_rel = freq_by_type.div(freq_by_type.sum(axis=0), axis=1)

    plt.figure()

    types2=[ 'plainte', 'porte_a_porte','enquete_entomo']
    for i, type_enquete in enumerate(types2):
        plt.bar(
            x + offsets[i],
            freq_rel[type_enquete],
            width=width,
            label=type_enquete
        )

    plt.xlabel("Nombre de lignes par triplet")
    plt.ylabel("Fréquence relative")
    plt.title(title_s)
    plt.xticks(x)
    plt.legend()
    plt.grid(axis="y")
    plt.show()

if LOAD_DONNEES_BATIMENTS_VF:
    table='donnees_batiments'
    schema='traitements'
    col_geometry='geom'
    
    gpd_batiments=obj_data_access_services.get_data_table_to_gdf(tablename=table,col_geometry=col_geometry,schema=schema)
    print(gpd_batiments.head(1))
    

if LOAD_RAW_FEATURES_VF:
    table='raw_features'
    schema='points_enquete'
    col_geometry='geometry'
    nb_maxi=-1

    gpd_raw_features=obj_data_access_services.get_data_table_to_gdf(tablename=table,col_geometry=col_geometry,schema=schema,nb_maxi=nb_maxi)
    print(f'RAW FEATURES:')
    print(f'->{len(gpd_raw_features)} ligne(s)')
    print(f'->col: \n{gpd_raw_features.columns}')
   
    #sys.exit()
    gpd_filtered_features=gpd_raw_features.copy()

    columns_to_drop = [
    'id_maison_b50_m',
    'id_batiment_b50_m',
    'surf_batiment_source_m2_b50_m']

    gpd_filtered_features = gpd_filtered_features.drop(
    columns=columns_to_drop,
    errors='ignore'  # évite une erreur si une colonne est absente
    )

    gpd_filtered_features.rename(columns={'id_batiment_b10_m':'id_batiment','surf_batiment_source_m2_b10_m':'surf_batiment_source_m2'},inplace=True)
    #
    gpd_filtered_features=gpd_filtered_features[gpd_filtered_features['visite_VF']==True]
    gpd_filtered_features=gpd_filtered_features[(gpd_filtered_features['contenant enterré']>=0) & (gpd_filtered_features['grand contenant']>=0) & (gpd_filtered_features['petit contenant']>=0)]
    #gpd_filtered_features=gpd_filtered_features[gpd_filtered_features['type_enquete'].isin(['plainte','enquete_entomo'])]
 
    print(f'\nFILTERED FEATURES:')
    print(f'->{len(gpd_filtered_features)} ligne(s)')
    print(f'->col: \n{gpd_filtered_features.columns}')

    print(gpd_filtered_features['type_enquete'].unique())
    
   
    print(gpd_filtered_features[['contenant enterré','grand contenant', 'petit contenant']].describe())

    print(gpd_filtered_features.columns)
    

    print("nb points par agent:")
    print(gpd_filtered_features['code_agent'].value_counts())

    
    def plot_boxplots_features(
        df,
        features,
        figsize=(8, 8),
        title=None,
        showfliers=True
    ):
        """
        Display vertical boxplots for selected integer features.

        Parameters
        ----------
        df : pandas.DataFrame or geopandas.GeoDataFrame
            Input dataframe.
        features : list of str
            Column names to plot.
        figsize : tuple, optional
            Figure size.
        title : str, optional
            Global title.
        showfliers : bool, optional
            Whether to display outliers.
        """

        fig, axes = plt.subplots(
            nrows=len(features),
            ncols=1,
            figsize=figsize,
            sharex=False
        )

        # Ensure axes is iterable even for a single feature
        if len(features) == 1:
            axes = [axes]

        for ax, feature in zip(axes, features):
            data = df[feature].dropna()

            ax.boxplot(
                data,
                vert=False,
                showfliers=showfliers
            )

            ax.set_title(feature)
            ax.set_xlabel("Value (integer count)")
            ax.grid(axis='x', linestyle='--', alpha=0.5)

        if title:
            fig.suptitle(title, fontsize=14)
            plt.subplots_adjust(top=0.92)

        plt.tight_layout()
        plt.show()


        # Liste des colonnes à afficher
    features = [
        'contenant enterré',
        'grand contenant',
        'petit contenant'
    ]



    visu_boxplots_vf=False
    visu_freq_p95_vf=False

    if visu_boxplots_vf:
        plot_boxplots_features(
            df=gpd_filtered_features,
            features=features,
            title="Distribution des contenants – Boxplots",
            showfliers=True
        )

    if visu_freq_p95_vf:
        #Création de la figure avec 3 sous-graphiques verticaux
        fig, axes = plt.subplots(len(features), 1, figsize=(8, 10))

        for ax, feature in zip(axes, features):
            data = gpd_filtered_features[feature]

            # Quantile 0.95
            q95 = data.quantile(0.95)
            max_bin = math.ceil(q95)

            # Bins entiers
            bins = np.arange(0, max_bin + 2)  # +2 pour inclure la dernière classe

            ax.hist(
                data[data <= max_bin],
                bins=bins,
                edgecolor='black',
                align='left'
                )

            ax.set_title(f"{feature} (jusqu’au 95e percentile)")
            ax.set_xlabel("Nombre (valeurs entières)")
            ax.set_ylabel("Fréquence")

            ax.set_xticks(bins[:-1])

        plt.tight_layout()
        plt.show()



    features_col=[
        'id_maison',
        'surf_batiment_source_m2','hauteur',
        'surf_buffer_m2_b10_m', 'surf_batiment_b10_m', 'surf_broussaille_b10_m',
        'surf_conifere_b10_m', 'surf_feuillu_b10_m', 'surf_pelouse_b10_m',
        'surf_piscine_b10_m', 'surf_serre_b10_m', 'surf_sol_nu_b10_m',
        'surf_surface_eau_b10_m', 'surf_zone_impermeable_b10_m',
        'surf_zone_permeable_b10_m', 'surf_buffer_m2_b50_m',
        'surf_batiment_b50_m', 'surf_broussaille_b50_m', 'surf_conifere_b50_m',
        'surf_feuillu_b50_m', 'surf_pelouse_b50_m', 'surf_piscine_b50_m',
        'surf_serre_b50_m', 'surf_sol_nu_b50_m', 'surf_surface_eau_b50_m',
        'surf_zone_impermeable_b50_m', 'surf_zone_permeable_b50_m']

    targets_col=['id_maison', 'contenant enterré',
        'grand contenant', 'petit contenant', 'terrasses sur plots',
        'nb_conteneurs_total' ]


    row_nb=len(gpd_filtered_features)
    print('Nb de valeurs 0:')
    for f in features:
        c0=len(gpd_filtered_features[gpd_filtered_features[f]==0])
        print(f'{f}: {c0} / {row_nb}')

    TEST_NULLS_VF=False
    if TEST_NULLS_VF:
        null_counts = gpd_filtered_features.isna().sum()
        null_percent = (
            gpd_filtered_features.isna().mean() * 100
        ).round(2)
        print(f'->valeurs nulles: {null_counts}')
        print(f'->valeurs nulles pc: {null_percent}')
        print(null_counts[null_counts > 0].sort_values(ascending=False))
        print(null_percent[null_percent > 0].sort_values(ascending=False))

    #CORRIGE HAUTEURS:
    DEFAULT_HEIGHT = 5.0

    height_effective = (
        gpd_filtered_features["hauteur"]
        .astype(float)
        .fillna(DEFAULT_HEIGHT)
    )
    gpd_filtered_features["hauteur_corrigee_m"]=height_effective

    features_col.remove('hauteur')
    features_col=features_col+["hauteur_corrigee_m"]


    TEST_ABC_VF=True
    if TEST_ABC_VF:
        service = prepa_data_services.BuildingsCountDataPreparationService(id_col="id_maison")

        dataset_obj = service.prepare_dataset(
            df=gpd_filtered_features,
            features_col=features_col,
            targets_col=targets_col,
            add_volume=True,
            make_ratios=True,
            add_log1p=True,
            target_cols_to_keep=["contenant enterré", "grand contenant", "petit contenant"],
        )

        #__________________________________________
        targets="contenant enterré", "grand contenant", "petit contenant"
        #__________________________________________

        X, Y, ids = dataset_obj.X, dataset_obj.Y, dataset_obj.ids

        # print(f'->colonnes x:\n {X.columns}')
        
        # print(f'->colonnes y:\n {Y.columns}')
        # sys.exit()


        #On peut enlever les surfaces brutes:
        # X = artifacts.X.copy()

        # drop_cols = [c for c in X.columns
        #             if (c.startswith("surf_") and ("_b10_m" in c or "_b50_m" in c)
        #                 and not c.startswith("surf_buffer_m2_"))]

        # X = X.drop(columns=drop_cols)

        # Build Model C feature set (Config A)
        X_all_features = dataset_obj.X.copy()

        keep_cols = []

        # C1: building-scale features
        keep_cols += [
            "surf_batiment_source_m2",
            "hauteur_corrigee_m",
            "volume_batiment",
            "log1p_surf_batiment_source_m2",
            "log1p_volume_batiment",
        ]

        # C2: buffer scale features
        keep_cols += [
            "surf_buffer_m2_b10_m",
            "surf_buffer_m2_b50_m",
            "log1p_surf_buffer_m2_b10_m",
            "log1p_surf_buffer_m2_b50_m",
        ]

        # C3: composition ratios
        ratio_cols = [c for c in X_all_features.columns if c.startswith("ratio_")]
        keep_cols += ratio_cols

        # Final X for model C
        X_C = X_all_features[keep_cols].copy()

        # Sanity checks (minimal)
        assert X_C.isna().sum().sum() == 0, "NaNs found in X_C"
        assert np.isfinite(X_C.to_numpy()).all(), "Non-finite values found in X_C"

        print("X_C shape:", X_C.shape)
        print("Number of ratio cols:", len(ratio_cols))


        VISU_CORREL_VF=False
        if VISU_CORREL_VF:
            
            features_col=dataset_obj.feature_names
            # print(features_col)
            # sys.exit()


            features_utiles=['surf_batiment_source_m2',
        'ratio_surf_batiment_b10_m', 'ratio_surf_broussaille_b10_m', 'ratio_surf_conifere_b10_m',
        'ratio_surf_feuillu_b10_m', 'ratio_surf_pelouse_b10_m', 'ratio_surf_piscine_b10_m',
        'ratio_surf_serre_b10_m', 'ratio_surf_sol_nu_b10_m', 'ratio_surf_surface_eau_b10_m',
        'ratio_surf_zone_impermeable_b10_m', 'ratio_surf_zone_permeable_b10_m','ratio_veg_b10_m', 'ratio_veg_b50_m']
            

            features_utiles= [
            "surf_batiment_source_m2",
            "hauteur_corrigee_m",
            "volume_batiment",
            "log1p_surf_batiment_source_m2",
            "log1p_volume_batiment",
            'ratio_veg_b10_m', 'ratio_veg_b50_m'
        ]


            

            dataset_target_names=dataset_obj.target_names
            print(dataset_target_names)
            #['contenant enterré', 'grand contenant', 'petit contenant']

            visu_scatters_vf=False
            if visu_scatters_vf:
                dataset_obj.plot_feature_target_scatter_matrix(target_col="contenant enterré", features_cols=features_utiles)

                dataset_obj.plot_feature_target_scatter_matrix(target_col="grand contenant", features_cols=features_utiles)

                dataset_obj.plot_feature_target_scatter_matrix(target_col="petit contenant", features_cols=features_utiles)
                plt.show()

            visu_spearman_correl_vf=True
            if visu_spearman_correl_vf:
                plot_vf=False
                features_utiles=None

                for target in targets:
                    
                    sp=dataset_obj.spearman_correlations(target_col=target, features_cols=features_utiles,plot=plot_vf)
                    print((f'\nSpearman correlation {target}:'))
                    print(sp)
              

                


            

            

           

        CLUSTERISATION_BATI_VF=False
        if CLUSTERISATION_BATI_VF:
            #print(dataset_obj.feature_names)
            col_utiles=['surf_batiment_source_m2', 'surf_buffer_m2_b10_m',
                             'hauteur_corrigee_m', 'volume_batiment',
                             'ratio_surf_batiment_b10_m', 'ratio_surf_broussaille_b10_m', 'ratio_surf_conifere_b10_m', 'ratio_surf_feuillu_b10_m',  'ratio_surf_surface_eau_b10_m', 'ratio_surf_zone_impermeable_b10_m', 'ratio_surf_zone_permeable_b10_m',
                             'ratio_surf_batiment_b50_m', 'ratio_surf_broussaille_b50_m', 'ratio_surf_conifere_b50_m', 'ratio_surf_feuillu_b50_m',  'ratio_surf_surface_eau_b50_m', 'ratio_surf_zone_impermeable_b50_m', 'ratio_surf_zone_permeable_b50_m',
                            ]
            data_x=dataset_obj.X[col_utiles]
            from sklearn.preprocessing import StandardScaler
            from sklearn.preprocessing import RobustScaler

            X_scaled = StandardScaler().fit_transform(data_x)
            #X_scaled = RobustScaler().fit_transform(data_x)

            print(type(X_scaled))
            df_scaled=pd.DataFrame(X_scaled,columns=data_x.columns)
            corr = df_scaled.corr(method="pearson")
            print(corr)


            from sklearn.decomposition import PCA

            pca = PCA()
            X_pca = pca.fit_transform(df_scaled)
            explained_variance = pca.explained_variance_ratio_
            print("Variance expliquée:")
            print(explained_variance)
            #
            plt.plot(range(1, len(explained_variance) + 1), explained_variance, marker="o")
            plt.xlabel("Composante")
            plt.ylabel("Variance expliquée")
            plt.title("Scree plot ACP")
            plt.show()

           

            def visu_acp( composante_1=0, composante_2=1):
                plt.figure(figsize=(6, 6))
                plt.scatter(X_pca[:, composante_1], X_pca[:, composante_2], s=5, alpha=0.4)
                plt.axhline(0, linewidth=0.5)
                plt.axvline(0, linewidth=0.5)
                plt.xlabel("PC1")
                plt.ylabel("PC2")
                plt.title(f"ACP – projection des individus (PC{composante_1+1}–PC{composante_2+1})")
                plt.show()

                plt.hexbin(X_pca[:, composante_1], X_pca[:, composante_2], gridsize=40)
                plt.xlabel(f"PC{composante_1+1}")
                plt.ylabel(f"PC{composante_2+1}")
                plt.title(f"ACP – densité des individus (PC{composante_1+1}–PC{composante_2+1})")
                plt.colorbar(label="Densité")
                plt.show()

                plt.scatter(X_pca[:, composante_2], X_pca[:, composante_1], s=5, cmap="viridis")
                plt.colorbar(label=f"PC{composante_1+1}")
                plt.show()

            visu_acp( composante_1=0, composante_2=1)
            visu_acp( composante_1=1, composante_2=2)



        #_______________________________________
        #
        #MODELES 
        #
        #_______________________________________

    

    features_utiles=dataset_obj.feature_names
    # features_utiles=['surf_batiment_source_m2',
    #     'ratio_surf_batiment_b10_m', 'ratio_surf_broussaille_b10_m', 'ratio_surf_conifere_b10_m',
    #     'ratio_surf_feuillu_b10_m', 'ratio_surf_pelouse_b10_m', 'ratio_surf_piscine_b10_m',
    #     'ratio_surf_serre_b10_m', 'ratio_surf_sol_nu_b10_m', 'ratio_surf_surface_eau_b10_m',
    #     'ratio_surf_zone_impermeable_b10_m', 'ratio_surf_zone_permeable_b10_m','ratio_veg_b10_m', 'ratio_veg_b50_m']
            
    

    NEW_MODELS_VF=False
    PREDICTION_MONO_VF=True

    lgbm_params = {
                    "objective": "poisson",
                    "metric": "poisson",
                    "learning_rate": 0.05,
                    "num_leaves": 31,
                    "min_data_in_leaf": 50,
                    "feature_fraction": 0.8,
                    "verbosity": -1,
                    "seed": 42
                }
    
    if NEW_MODELS_VF:
        if PREDICTION_MONO_VF:

            import numpy as np
            import pandas as pd

            from sklearn.model_selection import train_test_split
            from sklearn.metrics import mean_absolute_error, mean_squared_error, mean_poisson_deviance,root_mean_squared_error

            import lightgbm as lgb
            import statsmodels.api as sm

            def compare_count_models(
                dataset_obj,
                features,
                target,
                test_size=0.2,
                random_state=42
            ):
                # -------------------------
                # Train / test split
                # -------------------------
                X = dataset_obj.X[features]
                y = dataset_obj.Y[target]

                X_train, X_test, y_train, y_test = train_test_split(
                    X, y, test_size=test_size, random_state=random_state
                )

                results = []

                # =====================================================
                # 1) Mean baseline
                # =====================================================
                mean_pred = np.full_like(y_test, y_train.mean(), dtype=float)
                mean_pred = np.clip(mean_pred, 1e-6, None)

                results.append({
                    "model": "Mean_baseline",
                    "MAE": mean_absolute_error(y_test, mean_pred),
                    "RMSE": root_mean_squared_error(y_test, mean_pred),
                    "Poisson_deviance": mean_poisson_deviance(y_test, mean_pred)
                })

                # =====================================================
                # 2) Negative Binomial regression
                # =====================================================
                X_train_sm = sm.add_constant(X_train, has_constant="add")
                X_test_sm = sm.add_constant(X_test, has_constant="add")

                nb_model = sm.GLM(
                    y_train,
                    X_train_sm,
                    family=sm.families.NegativeBinomial()
                ).fit()

                nb_pred = nb_model.predict(X_test_sm)
                nb_pred = np.clip(nb_pred, 1e-6, None)

                results.append({
                    "model": "Negative_Binomial",
                    "MAE": mean_absolute_error(y_test, nb_pred),
                    "RMSE": root_mean_squared_error(y_test, nb_pred),
                    "Poisson_deviance": mean_poisson_deviance(y_test, nb_pred)
                })

                # =====================================================
                # 3) LightGBM Poisson
                # =====================================================
                lgb_train = lgb.Dataset(X_train, label=y_train)
                lgb_test = lgb.Dataset(X_test, label=y_test, reference=lgb_train)

               

                gbm = lgb.train(
                    lgbm_params,
                    lgb_train,
                    num_boost_round=500,
                    valid_sets=[lgb_test],
                    callbacks=[lgb.early_stopping(50, verbose=False)]
                )

                lgb_pred = gbm.predict(X_test)
                lgb_pred = np.clip(lgb_pred, 1e-6, None)

                results.append({
                    "model": "LightGBM_Poisson",
                    "MAE": mean_absolute_error(y_test, lgb_pred),
                    "RMSE": root_mean_squared_error(y_test, lgb_pred),
                    "Poisson_deviance": mean_poisson_deviance(y_test, lgb_pred)
                })

                df = pd.DataFrame(results)

                # -------------------------
                # Normalisation vs baseline
                # -------------------------
                baseline = df.loc[df["model"] == "Mean_baseline"].iloc[0]

                for metric in ["MAE", "RMSE", "Poisson_deviance"]:
                    df[f"{metric}_rel_impr"] = (baseline[metric] - df[metric]) / baseline[metric]
                    df[f"{metric}_rel_impr_pct"] = 100 * df[f"{metric}_rel_impr"]

                # tri selon la métrique principale
                return df.sort_values("Poisson_deviance")

        
            print("CALCUL MONO:")
            for target in targets:
                print(f'TARGET {target}:')    
                results=compare_count_models(dataset_obj, features=features_utiles, target=target)
                print(results)
                rep_out=r'C:\Users\aubin\ACTIONS2\Geo2I\Moustiques\Analyse_fichier_moustique'
                file_out=os.path.join(rep_out,f'model_mono_{target.replace(' ','_')}.csv')
                print(file_out)
                results.to_csv(file_out)

           

  



    baseline_B_features = [
        "log1p_surf_batiment_source_m2",
        "hauteur_corrigee_m",
        "log1p_volume_batiment",
        'log1p_surf_buffer_m2_b10_m', 'log1p_surf_buffer_m2_b50_m'
    ]



    pks=( 10, 20, 40, 60, 90)
    #pks=( 2,4,5,10)

    groups_regression = 40
    


    p_n_draws=1000
    coords = gpd_filtered_features.loc[X_C.index, ["x_visite", "y_visite"]].to_numpy()

    lim_graphes_correl={}
    lim_graphes_correl["contenant enterré"]=(0,140)
    lim_graphes_correl["grand contenant"]=(0,150)
    lim_graphes_correl["petit contenant"]=(0,450)


    if TEST_BASES_LINES_VF:
        for target in targets:        
            test_models.set_test_A_B(target,X,Y,avec_agreg=False)

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    def export_concat_table_to_excel_merged_headers(
        df: pd.DataFrame,
        filepath: str,
        sheet_name: str = "Summary",
        freeze_panes: str = "A3",
        auto_width: bool = True,
    ) -> None:
        """
        Export a DataFrame with 2-level MultiIndex columns to a single Excel sheet
        with merged top headers (level 0) and second-row subheaders (level 1).

        If df.columns is not a MultiIndex, it will still export, but without merges.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # -------------------------
        # Prepare headers
        # -------------------------
        if isinstance(df.columns, pd.MultiIndex) and df.columns.nlevels >= 2:
            col_level0 = [str(c[0]) if c[0] is not None else "" for c in df.columns]
            col_level1 = [str(c[1]) if c[1] is not None else "" for c in df.columns]
        else:
            col_level0 = [str(c) for c in df.columns]
            col_level1 = [""] * len(df.columns)

        n_cols = len(df.columns)

        # -------------------------
        # Write header rows
        # -------------------------
        # Row 1: level 0
        for j in range(n_cols):
            cell = ws.cell(row=1, column=j + 1, value=col_level0[j])
            cell.font = header_font
            cell.alignment = center

        # Row 2: level 1
        for j in range(n_cols):
            cell = ws.cell(row=2, column=j + 1, value=col_level1[j])
            cell.font = header_font
            cell.alignment = center

        # Merge contiguous identical labels in row 1
        j = 0
        while j < n_cols:
            label = col_level0[j]
            start = j
            end = j
            while end + 1 < n_cols and col_level0[end + 1] == label:
                end += 1
            if end > start and label != "":
                ws.merge_cells(start_row=1, start_column=start + 1, end_row=1, end_column=end + 1)
            j = end + 1

        # -------------------------
        # Write data
        # -------------------------
        for i, row in enumerate(df.itertuples(index=False), start=3):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # -------------------------
        # Styling / ergonomics
        # -------------------------
        ws.freeze_panes = freeze_panes

        # Center headers already; you can also align all cells left/center if desired.
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

        # Auto width (basic)
        if auto_width:
            for col_idx in range(1, n_cols + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, min(ws.max_row, 2000) + 1):  # cap scan for speed
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

        wb.save(filepath)





    def export_tables_to_excel(tables: dict[str, pd.DataFrame], filepath: str) -> None:
        """
        Export all target tables into ONE Excel sheet, concatenated, with merged headers
        if MultiIndex columns are present. Same call signature as before.
        """
        if not tables:
            raise ValueError("tables is empty")

        # --- concat tables (keep MultiIndex columns!)
        dfs = []
        for target_name, df in tables.items():
            df2 = df.copy()

            if isinstance(df2.columns, pd.MultiIndex):
                if ("Target", "") in df2.columns:
                    df2[("Target", "")] = df2[("Target", "")].fillna(target_name)
                else:
                    df2.insert(0, ("Target", ""), target_name)
            else:
                if "Target" in df2.columns:
                    df2["Target"] = df2["Target"].fillna(target_name)
                else:
                    df2.insert(0, "Target", target_name)

            dfs.append(df2)
            blank = pd.DataFrame([[np.nan] * df2.shape[1]], columns=df2.columns)
            dfs.append(blank)

        df_all = pd.concat(dfs, ignore_index=True)

        # --- write with openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "CV_ALL"

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # If MultiIndex columns -> 2 header rows. Otherwise 1.
        is_mi = isinstance(df_all.columns, pd.MultiIndex) and df_all.columns.nlevels >= 2

        if is_mi:
            metrics = [str(c[0]) if c[0] is not None else "" for c in df_all.columns]
            stats   = [str(c[1]) if c[1] is not None else "" for c in df_all.columns]

            # Row 1: metric (merge identical adjacent metrics)
            for j, v in enumerate(metrics, start=1):
                cell = ws.cell(row=1, column=j, value=v)
                cell.font = header_font
                cell.alignment = center

            # Row 2: stat (mean/std)
            for j, v in enumerate(stats, start=1):
                cell = ws.cell(row=2, column=j, value=v)
                cell.font = header_font
                cell.alignment = center

            # Merge row 1 blocks
            j = 0
            n_cols = len(metrics)
            while j < n_cols:
                lab = metrics[j]
                start = j
                end = j
                while end + 1 < n_cols and metrics[end + 1] == lab:
                    end += 1
                if lab != "" and end > start:
                    ws.merge_cells(start_row=1, start_column=start + 1, end_row=1, end_column=end + 1)
                j = end + 1

            data_start_row = 3

        else:
            # Single header row
            for j, v in enumerate(df_all.columns, start=1):
                cell = ws.cell(row=1, column=j, value=str(v))
                cell.font = header_font
                cell.alignment = center
            data_start_row = 2

        # Write data
        for i, row in enumerate(df_all.itertuples(index=False), start=data_start_row):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # Freeze panes
        ws.freeze_panes = f"A{data_start_row}"

        # Auto width (simple)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row_idx in range(1, min(ws.max_row, 1200) + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

        wb.save(filepath)


    def concat_tables_with_separators(tables: list[pd.DataFrame], sep_rows: int = 1) -> pd.DataFrame:
        """Concatenate multiple tables (same columns) and insert blank separator rows."""
        if not tables:
            raise ValueError("No tables provided.")
        cols = tables[0].columns
        out_parts = []
        for i, df in enumerate(tables):
            if not df.columns.equals(cols):
                raise ValueError("All tables must have the same columns to be concatenated.")
            out_parts.append(df)
            if i < len(tables) - 1 and sep_rows > 0:
                blank = pd.DataFrame([[np.nan] * len(cols)] * sep_rows, columns=cols)
                out_parts.append(blank)
        return pd.concat(out_parts, ignore_index=True)


    def export_concat_table_to_excel_with_section_headers(
        df: pd.DataFrame,
        filepath: str,
        sheet_name: str = "Summary",
        freeze_panes: str = "A4",
        auto_width: bool = True,
    ) -> None:
        """
        Export a DataFrame with MultiIndex columns (metric, stat) to Excel with:
        Row 1: section headers (your exact titles) with merged cells
        Row 2: metric headers (merged by metric)
        Row 3: subheaders (mean/std)
        Row 4+: data

        Expected conventions for metric names (level 0):
        - RSE_*                         -> section "RSE sur les valeurs sommées"
        - *_sum (Pearson_sum, MAE_sum…) -> section "MESURES QUALITE sur les VALEURS SOMMÉES"
        - *_meanGrp                     -> section "MESURES QUALITE sur les VALEURS MOYENNES"
        - N_groups                      -> section "Population"
        - plus left info columns like Model/Group_size/Target (non-section, kept as is)
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # ----- Extract columns
        if isinstance(df.columns, pd.MultiIndex) and df.columns.nlevels >= 2:
            metrics = [str(c[0]) if c[0] is not None else "" for c in df.columns]
            stats   = [str(c[1]) if c[1] is not None else "" for c in df.columns]
        else:
            # fallback: no multiindex
            metrics = [str(c) for c in df.columns]
            stats = [""] * len(df.columns)

        n_cols = len(df.columns)

        # ----- Identify sections (your exact titles)
        # Left “info” columns are those not matching known patterns
        def _section_for_metric(m: str) -> str:
            if m.startswith("RSE_"):
                return "RSE sur les valeurs sommées"
            if m.endswith("_sum"):
                return "MESURES QUALITE sur les VALEURS SOMMÉES"
            if m.endswith("_meanGrp"):
                return "MESURES QUALITE sur les VALEURS MOYENNES"
            if m == "N_groups":
                return "Population"
            # anything else (Model, Group_size, Target, etc.)
            return ""

        sections = [_section_for_metric(m) for m in metrics]

        # ----- Write header rows
        # Row 1: section headers
        for j in range(n_cols):
            c = ws.cell(row=1, column=j + 1, value=sections[j])
            c.font = header_font
            c.alignment = center

        # Row 2: metric headers (level 0)
        for j in range(n_cols):
            c = ws.cell(row=2, column=j + 1, value=metrics[j])
            c.font = header_font
            c.alignment = center

        # Row 3: subheaders (mean/std)
        for j in range(n_cols):
            c = ws.cell(row=3, column=j + 1, value=stats[j])
            c.font = header_font
            c.alignment = center

        # ----- Merge contiguous identical labels on row 1 (sections)
        j = 0
        while j < n_cols:
            label = sections[j]
            start = j
            end = j
            while end + 1 < n_cols and sections[end + 1] == label:
                end += 1
            # Merge only if label non-empty and spans > 1 col
            if label != "" and end > start:
                ws.merge_cells(start_row=1, start_column=start + 1, end_row=1, end_column=end + 1)
            j = end + 1

        # ----- Merge contiguous identical labels on row 2 (metrics)
        j = 0
        while j < n_cols:
            label = metrics[j]
            start = j
            end = j
            while end + 1 < n_cols and metrics[end + 1] == label:
                end += 1
            if label != "" and end > start:
                ws.merge_cells(start_row=2, start_column=start + 1, end_row=2, end_column=end + 1)
            j = end + 1

        # ----- Write data starting row 4
        for i, row in enumerate(df.itertuples(index=False), start=4):
            for j, val in enumerate(row, start=1):
                ws.cell(row=i, column=j, value=val)

        # ----- Cosmetic
        ws.freeze_panes = freeze_panes
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18

        # Auto width (basic, capped)
        if auto_width:
            for col_idx in range(1, n_cols + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, min(ws.max_row, 1500) + 1):
                    v = ws.cell(row=row_idx, column=col_idx).value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))
                ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 45)

        wb.save(filepath)


    def flatten_multiindex_columns(df: pd.DataFrame, sep: str = "__") -> pd.DataFrame:
        """
        Flatten MultiIndex columns into single-level columns.
        Example: ('RSE_mean','mean') -> 'RSE_mean__mean'
        """
        out = df.copy()
        if isinstance(out.columns, pd.MultiIndex):
            out.columns = [
                sep.join([str(x) for x in col if x is not None and str(x) != ""])
                for col in out.columns.to_flat_index()
            ]
        return out


    def plot_summaries_by_target(
        summaries_by_target: dict,
        metrics: list[str] = ("RSE_median", "RSE_p95"),
        stat: str = "mean",                   # "mean" or "std"
        model_order: list[str] | None = None, # optional ordering
        group_order: list[int] | None = None, # optional ordering
        grouping_label: str | None = None,    # "knn" or "random" (for titles)
        mode_label: str | None = None,        # "sum" or "mean" (for titles)
        round_digits: int = 3,
        annotate: bool = True,
        figsize_per_cell: tuple[float, float] = (6.2, 4.2),
    ):
        """
        Plot CV summary curves from your 'summary_spatial' tables.

        Parameters
        ----------
        summaries_by_target : dict[target -> summary_df]
            Each value is what you call `summary_spatial` (output of build_readable_cv_table),
            either with MultiIndex columns (metric, stat) OR flattened columns like "RSE_median__mean".
        metrics : list[str]
            Metric names exactly as used in summary (e.g. "RSE_median", "RSE_p95", "MAE_sum", "MAE_meanGrp"...).
            NOTE: RSE metrics are independent of sum/mean; sum/mean-specific metrics must be chosen accordingly.
        stat : "mean" or "std"
            Which fold-aggregation statistic to plot.
        model_order, group_order : optional
            If provided, enforce ordering of model and group sizes.
        grouping_label : optional
            String used only for title. Example: "knn" or "random".
        mode_label : optional
            String used only for title. Example: "sum" or "mean".
        annotate : bool
            If True, prints numeric values on each point.
        """

        if isinstance(metrics, str):
            metrics = [metrics]

        # --- Helpers to read values regardless of column style (MultiIndex vs flattened)
        def _has_multiindex_columns(df: pd.DataFrame) -> bool:
            return isinstance(df.columns, pd.MultiIndex)

        def _get_series(df: pd.DataFrame, metric: str, stat: str) -> pd.Series:
            if _has_multiindex_columns(df):
                # columns like ("RSE_median", "mean")
                if (metric, stat) not in df.columns:
                    raise KeyError(f"Missing column {(metric, stat)} in MultiIndex columns.")
                return df[(metric, stat)]
            else:
                # columns like "RSE_median__mean"
                col = f"{metric}__{stat}"
                if col not in df.columns:
                    raise KeyError(f"Missing column '{col}' in flattened columns.")
                return df[col]

        # --- Prepare layout
        targets = list(summaries_by_target.keys())
        n_targets = len(targets)
        n_metrics = len(metrics)

        fig_w = figsize_per_cell[0] * n_metrics
        fig_h = figsize_per_cell[1] * n_targets
        fig, axes = plt.subplots(nrows=n_targets, ncols=n_metrics, figsize=(fig_w, fig_h), squeeze=False,constrained_layout=True)

        # --- Plot each target row
        for i, target in enumerate(targets):
            df = summaries_by_target[target].copy()

            # Basic columns (your summary has these)
            # Works both for MultiIndex and flattened columns
            if "Model" not in df.columns or "Group_size" not in df.columns:
                raise ValueError(
                    f"Summary for target '{target}' must contain columns 'Model' and 'Group_size'."
                )

            # Order model and group if requested
            if model_order is not None:
                df["Model"] = pd.Categorical(df["Model"], categories=model_order, ordered=True)
            if group_order is not None:
                df["Group_size"] = pd.Categorical(df["Group_size"], categories=group_order, ordered=True)

            df = df.sort_values(["Model", "Group_size"])

            for j, metric in enumerate(metrics):
                ax = axes[i, j]

                # One curve per model
                models = model_order if model_order is not None else list(df["Model"].unique())

                for m in models:
                    dfi = df[df["Model"] == m].copy()
                    if dfi.empty:
                        continue

                    x = dfi["Group_size"].astype(float).to_numpy()
                    y = _get_series(dfi, metric, stat).astype(float).to_numpy()

                    ax.plot(x, y, marker="o", linewidth=2, label=str(m))

                    if annotate:
                        for xv, yv in zip(x, y):
                            if np.isfinite(yv):
                                ax.annotate(
                                    f"{yv:.{round_digits}f}",
                                    (xv, yv),
                                    textcoords="offset points",
                                    xytext=(0, 7),
                                    ha="center",
                                    fontsize=9
                                )

                ax.set_xlabel("Group size (k)")
                ax.set_ylabel(f"{metric} ({stat})")
                ax.grid(True, alpha=0.2)

                # Title per cell
                title_parts = [str(target), metric]
                if grouping_label:
                    title_parts.append(f"grouping={grouping_label}")
                if mode_label:
                    title_parts.append(f"mode={mode_label}")
                ax.set_title(" | ".join(title_parts))

                # legend only on first row / first metric (avoid repetition)
                if i == 0 and j == 0:
                    ax.legend(loc="best", frameon=True)

        # Global title
        suptitle = "CV summary curves"
        if grouping_label or mode_label:
            extras = []
            if grouping_label:
                extras.append(f"grouping={grouping_label}")
            if mode_label:
                extras.append(f"mode={mode_label}")
            suptitle += " — " + ", ".join(extras)

        fig.suptitle(suptitle, y=1.02, fontsize=14)

        #plt.tight_layout(rect=[0, 0, 1, 0.95])  # laisse 5% en haut pour le titre

        plt.show()
       
        return fig, axes


    if  TEST_AGGREG_ABC_KNN_VF:
        Test_aggreg_ABC_KNN_summaries={}

        for target in targets:  
            print(f'Target: {target}')      
            y = dataset_obj.Y[target]
            max_groups=None
            #c
            #mode_groupement="random"
            mode_groupement="knn"
            #

            spatial_results =comparison_regression_models_services.cv_spatial_knn_protocol_ABC(
                X_B=X_C,
                X_C=X_C,
                coords=coords,
                y=y,
                baseline_B_features=baseline_B_features,
                k_groups=pks,
                max_groups=max_groups,
                grouping=mode_groupement
            )

            summary_spatial = comparison_regression_models_services.build_readable_cv_table(
            spatial_results,
            target_name=target,
            model_order=["A_mean", "B_neg_bin", "C_lgbm_poisson"],
            group_order=pks,
            round_digits=3
             )

            Test_aggreg_ABC_KNN_summaries[target]=summary_spatial

            print(f'\nSummary: {target}:')

            print(summary_spatial)
            print(f'\nFin summary: {target}:')
 
            

            # rep_out=r'C:\Users\aubin\ACTIONS2\Geo2I\Moustiques\Analyse_fichier_moustique_v2'
            # file_out=os.path.join(rep_out,f'knn_{target.replace(' ','_')}.csv')
            # summary_spatial.to_csv(file_out)

        
            # if TEST_VISU_COURBES_REGRESSION_VF:
            #     #targets="contenant enterré", "grand contenant", "petit contenant"
            
            #     def scatters_results(target, k,min_max=None):
            #         max_groups_per_fold=100
            #         #
            #         modelC_factory =lambda: modeles_services_regression.ModelCPoissonLGBM(params=None, random_state=42)
            #         #
            #         true_sums_C, pred_sums_C = comparison_regression_models_services.cv_collect_group_sums_modelC(
            #         X_C=X_C,
            #         coords=coords,
            #         y=y,
            #         modelC_factory=modelC_factory,
            #         k=k,
            #         max_groups_per_fold=max_groups_per_fold
            #     )

            #         visu.plot_true_vs_pred_sector_sums(
            #             true_sums_C, pred_sums_C,
            #             k=k,
            #             min_max=min_max,
            #             title=f"{target} — Totaux par pseudo-secteur k={k} (CV, modèle C)"
            #         )

            #     use_min_max_vf=False
            #     if use_min_max_vf:
            #         min_max=lim_graphes_correl[target]
            #     else:
            #         min_max=None
            #     scatters_results(target=target,min_max=min_max, k=groups_regression)

        export_excel_vf=True
        if export_excel_vf:
            rep_out=r'C:\Users\aubin\ACTIONS2\Geo2I\Moustiques\Analyse_fichier_moustique'
            file_out=os.path.join(rep_out,f'cv_knn_summary_{mode_groupement}.xlsx')

            export_tables_to_excel(Test_aggreg_ABC_KNN_summaries, file_out)
            print(f'knn summary exporté vers:\n {file_out}')

       
        
    
    
    def list_available_metrics(summary_df: pd.DataFrame) -> list[str]:
        """
        Return the list of available metric names in a summary table
        with MultiIndex columns (metric, stat).
        Keeps only numeric metric blocks (i.e. those that have 'mean'/'std').
        """
        if not isinstance(summary_df.columns, pd.MultiIndex):
            # flattened case: take prefix before "__"
            return sorted({c.split("__")[0] for c in summary_df.columns if "__" in c})

        lvl0 = summary_df.columns.get_level_values(0)
        lvl1 = summary_df.columns.get_level_values(1)

        # metrics are those having mean/std blocks (exclude the metadata columns with stat "")
        metrics = sorted({
            m for m, s in zip(lvl0, lvl1)
            if s in ("mean", "std") and m not in ("Model", "Group_size", "Target", "Grouping")
        })
        return metrics

    def default_metrics_for_mode(mode: str, kind: str = "quality") -> list[str]:
        """
        mode: "sum" or "mean"
        kind:
        - "quality": MAE/RMSE + corr
        - "bias": Bias + RelBias
        - "rse": RSE_median + RSE_p95 (indépendant du mode)
        """
        mode = mode.lower().strip()
        if kind == "rse":
            return ["RSE_median", "RSE_p95"]

        if mode == "sum":
            if kind == "quality":
                return ["Pearson_sum", "Spearman_sum", "MAE_sum", "RMSE_sum"]
            if kind == "bias":
                return ["Bias_sum", "RelBias_sum"]

        if mode == "mean":
            if kind == "quality":
                return ["Pearson_meanGrp", "Spearman_meanGrp", "MAE_meanGrp", "RMSE_meanGrp"]
            if kind == "bias":
                return ["Bias_meanGrp", "RelBias_meanGrp"]

        raise ValueError("mode must be 'sum' or 'mean', and kind in {'quality','bias','rse'}")



    if TEST_VISU_COURBES_AGREG_vf:
        stat="mean"  #"std"
       
        visu_indicateurs_dispo_vf=False
        if visu_indicateurs_dispo_vf:
            first_target = next(iter(Test_aggreg_ABC_KNN_summaries))
            summary0 = Test_aggreg_ABC_KNN_summaries[first_target]

            print(list_available_metrics(summary0))
        #['Bias_meanGrp', 'Bias_sum', 
        # 'MAE_meanGrp', 'MAE_sum', 
        # 'N_groups',
        #  'Pearson_meanGrp', 'Pearson_sum'
        # , 'RMSE_meanGrp', 'RMSE_sum',
        #  'RSE_mean', 'RSE_median', 'RSE_p90', 'RSE_p95',
        #  'RelBias_meanGrp', 'RelBias_sum',
        #  'Spearman_meanGrp', 'Spearman_sum']

        else:  #sinon...action!
            mode_label="moyenne"
            fig, axes=plot_summaries_by_target(
                summaries_by_target=Test_aggreg_ABC_KNN_summaries,
                metrics=["RSE_median", "RSE_p95"],
                stat=stat,
                grouping_label=mode_groupement,     # ou "random"
                mode_label=mode_label,              # juste pour info dans le titre (RSE s’applique aux deux)
                round_digits=3,
                annotate=True
            )
            plt.close(fig)

            mode_label="moyenne"
            fig, axes=plot_summaries_by_target(
                summaries_by_target=Test_aggreg_ABC_KNN_summaries,
                metrics=["MAE_meanGrp", "RMSE_meanGrp"],
                stat=stat,
                grouping_label=mode_groupement,     # ou "random"
                mode_label=mode_label,              # juste pour info dans le titre (RSE s’applique aux deux)
                round_digits=3,
                annotate=True
            )
            plt.close(fig)

            
            mode_label="moyenne"
            fig, axes=plot_summaries_by_target(
                summaries_by_target=Test_aggreg_ABC_KNN_summaries,
                metrics=["Pearson_meanGrp", "Spearman_meanGrp"],
                stat=stat,
                grouping_label=mode_groupement,     # ou "random"
                mode_label=mode_label,              # juste pour info dans le titre (RSE s’applique aux deux)
                round_digits=3,
                annotate=True
            )
            plt.close(fig)

            
            mode_label="cumul"
            fig, axes=plot_summaries_by_target(
                summaries_by_target=Test_aggreg_ABC_KNN_summaries,
                metrics=["Pearson_sum", "Spearman_sum"],
                stat=stat,
                grouping_label=mode_groupement,     # ou "random"
                mode_label=mode_label,              # juste pour info dans le titre (RSE s’applique aux deux)
                round_digits=3,
                annotate=True
            )
            plt.close(fig)



    if TEST_VISU_SCATTERS_par_GROUPE_v2_vf:
        k_list=[10, 50, 90]
        #mode="sum" # ou "mean"
        mode="mean"
        #grouping_mode="random"
        grouping_mode="knn"


        for target in targets:
            fig, axes =visu. plot_lgbm_true_vs_pred_by_group_size(
                dataset_obj=dataset_obj,
                coords=coords,                      # np.array (n,2) aligné avec dataset_obj
                target_col=target,
                features_cols=features_utiles,
                k_list=k_list,
                mode=mode,                            # ou "mean"
                max_groups=600,
                random_state=42,
                lgbm_params=lgbm_params,
                grouping=grouping_mode
            )
            plt.show()


    if TEST_APPROCHE_par_CLASSES_VF:
        #targets="contenant enterré", "grand contenant", "petit contenant"
        target = "contenant enterré"

        #methodes="quantile", "thresholds", "balanced_integers"
        methode="balanced_integers"
        n_classes=10
        creer_une_classe_specifique_pour_zero_vf=True

        #
        binning_service =calcul_classes_services.TargetBinningService()
        #
        y=dataset_obj.Y[target]
        binning_spec=calcul_classes_services.BinningSpec(method=methode, n_classes=n_classes, zero_as_own_class=creer_une_classe_specifique_pour_zero_vf)
        
        binning_obj = binning_service.build_classes(y,spec=binning_spec,)
        #=>
        y_class = binning_obj.y_class
        bin_log = binning_obj.log
        #
        #->Bornes par classe:
        bornes=bin_log['edges']
        class_bounds = [
        {
            "class_id": i,
            "lower_bound": low,
            "upper_bound": high
        }
        for i, (low, high) in enumerate(zip(bornes[:-1], bornes[1:]))
        ]
        #
        nb_obj_classes=sum([v for v in bin_log["class_counts"].values() ])
        #
        print('\n___________________________________')
        print(binning_obj.column_name)
        print(f'nb indiv par classe:{bin_log["class_counts"]}')
        
        print(f'nb classes: {len(bin_log["class_counts"])} vs attendu: {n_classes}')
        nb_obj_classes=sum([v for v in bin_log["class_counts"].values() ])
        print(f'nb objets classés: {nb_obj_classes}')
        print(f'bornes: {bornes}')
        print(f'\nlimites de classes:')
        
        for b in class_bounds:
            print(f'{b['class_id']}: {b['lower_bound']} -> {b['upper_bound']}')
        print('___________________________________')


        print(f'Colonne classe:')
        class_col_name=f'Class_{target}_{methode[0:2]}{n_classes}'
        print(class_col_name)
        #
        dataset_obj.X[class_col_name]=y_class


        #COMPARAISON des MODELES:
        regressor =modeles_services_regression.ModelCPoissonLGBM()

        y_float = dataset_obj.Y[target]  

        models = [            
            classif_baselines.BaselineMajorityClass(),
            classif_baselines.BaselineStratifiedRandom(random_state=42),
            RegressionToClassBaseline(
                regressor=regressor,
                binning_service=binning_service,
                binning_spec=binning_spec,
                y_continuous=y_float,
            ),
            classif_models.ModelLGBMClassifier(random_state=42),
            classif_models.ModelLogRegMultinomial(random_state=42)

        ]

        X_all_features = dataset_obj.X.copy()

        keep_cols = []

        # Test B columns
        keep_cols += [
            "surf_batiment_source_m2",
            "hauteur_corrigee_m",
            "volume_batiment",
            "log1p_surf_batiment_source_m2",
            "log1p_volume_batiment",
        ]
        X_B = X_all_features[keep_cols].copy()

        # C2: buffer scale features
        keep_cols += [
            "surf_buffer_m2_b10_m",
            "surf_buffer_m2_b50_m",
            "log1p_surf_buffer_m2_b10_m",
            "log1p_surf_buffer_m2_b50_m",
        ]

        # C3: composition ratios
        ratio_cols = [c for c in X_all_features.columns if c.startswith("ratio_")]
        keep_cols += ratio_cols

        # Final X for model C
        X_C = X_all_features[keep_cols].copy()

        # Sanity checks (minimal)
        assert X_C.isna().sum().sum() == 0, "NaNs found in X_C"
        assert np.isfinite(X_C.to_numpy()).all(), "Non-finite values found in X_C"

        print("X_C shape:", X_C.shape)
        print("Number of ratio cols:", len(ratio_cols))


        y_cls=X_all_features[class_col_name]


        #TEST X_B
        TEST_CLASSIF_X_B_FV=False
        if TEST_CLASSIF_X_B_FV:
            classModelComparisonService=comparison_classification_models_services.ClassModelComparisonService()
            res_XB = classModelComparisonService.compare(
            X=X_B,
            y_class=y_cls,
            models=models,
            )
            print(f'\n{target} X_B:')
            print(res_XB.summary)

        #TEST X_C
        TEST_CLASSIF_X_C_FV=True
        if TEST_CLASSIF_X_C_FV:
            classModelComparisonService=comparison_classification_models_services.ClassModelComparisonService()
            res_XC = classModelComparisonService.compare(
            X=X_C,
            y_class=y_cls,
            models=models,
            )
            print(f'\n{target} X_C:')
            print(res_XC.summary)

    if TEST_APPROCHE_GROUPEE_VF==True:
        #print(gpd_filtered_features.columns)
        #'contenant enterré','grand contenant', 'petit contenant'
        use_log1p_vf=True
        limites={}
        limites['grand']=(0,50)
        limites['petit']=(0,50)
        limites['enterré']=(0,50)


        cloud_visu.plot_container_scatters(gpd_filtered_features, use_log1p_vf=use_log1p_vf,limites=limites)
        
        conditionnal_histo_vf=False
        if conditionnal_histo_vf:
            targets='contenant enterré','grand contenant', 'petit contenant'
            for target_col in targets:
                for condition_col in targets:
                    if condition_col==target_col:
                        continue
                    cloud_visu.plot_conditional_histograms(gpd_filtered_features,target_col=target_col,condition_col=condition_col)

        print(cloud_visu.compute_spearman_correlations(gpd_filtered_features))

    





